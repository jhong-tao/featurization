#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
==================================================
@Project -> File   ：xxq -> main.py
@IDE    ：PyCharm
@Author ：jhong.tao
@Date   ：2021/6/7 10:34
@Desc   ：
==================================================
"""
from dateutil.parser import parse

import openpyxl
import pandas as pd


def statistics(path_log, path_user, path_out, column_name):
    log_sheet = pd.read_csv(path_log)
    user_sheet = pd.read_excel(path_user, sheet_name='Sheet1')
    df_log = pd.DataFrame(log_sheet)
    df_user = pd.DataFrame(user_sheet)
    # df_log['时间'] = df_log["时间"].map(lambda t: pd.to_datetime(t+":00", format='%Y年%m月%d日 %H:%M:%S'))
    df_log['时间'] = df_log["时间"].map(lambda t: t.replace('年', '-').replace('月', '-').replace('日', '').replace(' ', '/')+':00')
    # df_log['学号'] = df_log['用户全名'].map(lambda s: ''.join(re.findall("\d", s)))
    df_user['fullname'] = df_user['学号'].map(str)+df_user['姓名'].map(lambda str: str.replace("·", ""))
    wb_one_log = openpyxl.Workbook()
    wb_one_re = openpyxl.Workbook()
    ws_re = wb_one_re.create_sheet(title="统计结果")
    ws_re.append(['姓名', column_name, '持续时间'])
    for name in df_user['fullname']:
        df_one = df_log.loc[lambda df:df['用户全名'] == name, ]  # str(filter(str.isdigit, name))
        ws_log = wb_one_log.create_sheet(title=name)

        if df_one.shape[0] > 0:
            t1 = df_one['时间'][0:-1].values
            t2 = df_one['时间'][1:].values
            time_sum = parse(t1[0])-parse(t1[0])
            for j in range(t1.size):
                if (parse(t1[j])-parse(t2[j])).seconds < (25*60):
                    time_sum = time_sum + parse(t1[j])-parse(t2[j])
            num_col_name = df_one.loc[lambda df:df['事件名称'] == column_name, ].shape[0]
            ws_re.append([name, num_col_name, time_sum])
        else:
            ws_re.append([name, '0-请检查选课名单是否是名字错误', '0:0:0'])
        ws_log.append(list(df_one.columns.values))
        for index, row in df_one.iterrows():
            ws_log.append(row.tolist())

        wb_one_log.save(path_out + "one_log.xlsx")
    wb_one_re.save(path_out + "one_re.xlsx")


if __name__ == "__main__":
    path_log = "./data/input/logs_计算机网络21秋_日志.csv"
    path_user = "./data/input/计算机网络选课名单.xlsx"
    path_out = "./data/output/"
    column_name = "查看了课程"

    statistics(path_log, path_user, path_out, column_name)
